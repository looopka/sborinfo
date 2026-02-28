import os
import asyncio
from datetime import datetime
from functools import partial

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot.config import settings
from bot.utils.helpers import sanitize_folder_name


HEADERS = ["Кто загрузил", "Username", "Дата и время", "Комментарий", "Имя файла", "Ссылка"]
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _create_report_sync(local_path: str, rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Загрузки"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row in rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    col_widths = [25, 18, 20, 35, 35, 50]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"
    wb.save(local_path)


def _append_row_sync(local_path: str, row: list):
    if os.path.exists(local_path):
        wb = load_workbook(local_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Загрузки"
        ws.append(HEADERS)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        col_widths = [25, 18, 20, 35, 35, 50]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 25
        ws.freeze_panes = "A2"

    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(local_path)


async def update_event_report(
    event_name: str,
    upload_data: dict,
    temp_dir: str,
) -> str:
    """
    Скачивает текущий report.xlsx события (если есть), добавляет строку, загружает обратно.
    Возвращает путь к файлу на Яндекс Диске.
    """
    from bot.services import yandex_disk

    safe_event = sanitize_folder_name(event_name)
    root = settings.yandex_disk_root
    remote_report_path = f"{root}/{safe_event}/report.xlsx"
    local_report_path = os.path.join(temp_dir, f"report_{safe_event}.xlsx")

    # Скачиваем существующий отчёт
    if await yandex_disk.exists(remote_report_path):
        await yandex_disk.download_file(remote_report_path, local_report_path)

    # Формируем строку
    full_name_parts = []
    if upload_data.get("first_name"):
        full_name_parts.append(upload_data["first_name"])
    if upload_data.get("last_name"):
        full_name_parts.append(upload_data["last_name"])
    full_name = " ".join(full_name_parts) or "Неизвестно"
    username = f"@{upload_data['username']}" if upload_data.get("username") else "—"
    uploaded_at = upload_data.get("uploaded_at", datetime.now().strftime("%Y-%m-%d %H:%M"))
    comment = upload_data.get("comment") or "—"
    file_name = upload_data.get("file_name", "")
    link = upload_data.get("yandex_link", "")

    row = [full_name, username, uploaded_at, comment, file_name, link]

    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, partial(_append_row_sync, local_report_path, row))

    # Загружаем обратно на Яндекс Диск
    await yandex_disk.upload_file(local_report_path, remote_report_path)

    # Убираем временный файл
    if os.path.exists(local_report_path):
        os.remove(local_report_path)

    return remote_report_path
